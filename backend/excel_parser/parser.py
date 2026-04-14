"""Parse seller-sprite / Jungle Scout Excel exports into unified product & review dicts,
plus analysis sheets (market, competitor, profit, keyword, trend, conclusion).

Optimised: iter_rows() for streaming reads, ThreadPoolExecutor for parallel file parsing."""
import os
import re
import tempfile
import warnings
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from pydantic import ValidationError

from .schemas import Product

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

MAX_ANALYSIS_ROWS = 500  # cap for market/keyword analysis sheets

# ──────────────────────────────────────────────
# Column name normalisation (Chinese & English)
# ──────────────────────────────────────────────
_COLUMN_MAP: Dict[str, str] = {
    "ASIN": "asin", "品牌": "brand", "商品标题": "title", "标题(翻译)": "title_cn",
    "产品卖点": "bullet_points", "产品卖点(翻译)": "bullet_points_cn",
    "价格($)": "price", "Prime价格($)": "prime_price",
    "月销量": "monthly_sales", "月销售额($)": "monthly_revenue",
    "评分数": "ratings_total", "月新增\n评分数": "monthly_new_reviews",
    "评分": "rating", "留评率": "review_rate",
    "FBA($)": "fba_fee", "毛利率": "margin",
    "大类目": "main_category", "大类BSR": "main_bsr",
    "小类目": "sub_category", "小类BSR": "sub_bsr",
    "类目路径": "category_path", "上架时间": "launch_date", "上架天数": "days_on_market",
    "配送方式": "fulfillment", "卖家所属地": "seller_location",
    "卖家数": "seller_count", "Buybox卖家": "buybox_seller",
    "变体数": "variation_count", "Q&A数": "qa_count", "Coupon": "coupon",
    "A+页面": "a_plus", "重量": "weight", "体积": "volume",
    "商品详情页链接": "url", "商品主图": "image_url",
    "父ASIN": "parent_asin", "SKU": "sku", "详细参数": "product_overview",
    # English
    "Brand": "brand", "Product Title": "title", "Bullet Points": "bullet_points",
    "Price($)": "price", "Monthly Sales": "monthly_sales",
    "Monthly revenue($)": "monthly_revenue", "Monthly Sales growth": "monthly_sales_growth",
    "Rating": "rating", "Reviews": "ratings_total",
    "Category": "main_category", "Category BSR": "main_bsr",
    "Sub-Category": "sub_category", "Sub-Category BSR": "sub_bsr",
    "Category Path": "category_path", "Image URL": "image_url", "URL": "url",
    "ParentASIN": "parent_asin", "Product Overview": "product_overview",
    "Brand URL": "brand_url", "Variations": "variation_count",
    "Variation Sold": "variation_sold", "Variation revenue($)": "variation_revenue",
    "Image": "image", "#": "rank", "LQS": "lqs",
    # Category-analysis template variants (e.g., LED工作灯-类目分析.xlsx → BSR竞品数据 sheet)
    "产品标题（完整）": "title",          # full-width parens (as used in template)
    "产品标题(完整)": "title",            # half-width fallback
    "产品卖点摘要": "bullet_points",
    "评论数": "ratings_total",
    "月均销量": "monthly_sales",
    "月均销售额($)": "monthly_revenue",
    "月均销售额": "monthly_revenue",
    # BSR竞品数据 sheet 扩展列名
    "BSR小类排名": "sub_bsr",
    "BSR大类排名": "main_bsr",
    "产品类型": "product_type",
    "供电方式": "power_source",
    "亮度(LM)": "brightness",
    "防水等级": "waterproof",
    "卖家信息": "seller_info",
    "类型细分": "product_subtype",
    "色温(K)": "color_temp",
    "电池容量/电压": "battery_spec",
    "序号": "rank",
    "Prime": "prime",
}

_REVIEW_COLUMN_MAP: Dict[str, str] = {
    "ASIN": "asin", "标题": "title", "标题(翻译)": "title_cn",
    "内容": "body", "内容(翻译)": "body_cn",
    "VP评论": "verified_purchase", "Vine Voice评论": "vine_voice",
    "型号": "variant", "星级": "rating", "赞同数": "helpful_count",
    "评论链接": "review_url", "评论人": "reviewer", "所属国家": "country",
    "Title": "title", "Content": "body", "Rating": "rating",
    "Verified Purchase": "verified_purchase",
}

# Pre-build lowercase lookup for faster matching
_COL_LOWER = {k.lower(): v for k, v in _COLUMN_MAP.items()}
_REV_LOWER = {k.lower(): v for k, v in _REVIEW_COLUMN_MAP.items()}


def _normalise_header(raw: str) -> str:
    if not raw:
        return ""
    raw = str(raw).strip()
    if raw in _COLUMN_MAP:
        return _COLUMN_MAP[raw]
    return _COL_LOWER.get(raw.lower(), raw)


def _normalise_review_header(raw: str) -> str:
    if not raw:
        return ""
    raw = str(raw).strip()
    if raw in _REVIEW_COLUMN_MAP:
        return _REVIEW_COLUMN_MAP[raw]
    return _REV_LOWER.get(raw.lower(), raw)


def _safe_str(v) -> str:
    return str(v).strip() if v is not None else ""


def _safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        s = str(v).replace(",", "").replace("$", "").replace("%", "").replace("#", "").strip()
        return float(s)
    except (ValueError, TypeError):
        return None


# ──────────────────────────────────────────────
# Header detection (iter_rows based)
# ──────────────────────────────────────────────
def _get_headers(ws, max_scan: int = 5) -> Tuple[int, List[str]]:
    """Find the header row using iter_rows for speed."""
    row_idx = 0
    for row in ws.iter_rows(min_row=1, max_row=max_scan, max_col=min(ws.max_column or 1, 100), values_only=True):
        row_idx += 1
        strs = [str(c).strip() if c is not None else "" for c in row]
        non_empty = [s for s in strs if s]
        if "ASIN" in strs or "#" in strs:
            return row_idx, strs
        if len(non_empty) >= 5:
            return row_idx, strs
    return 1, []


def _header_text(headers: List[str]) -> str:
    return " ".join(h.lower() for h in headers if h)


# ──────────────────────────────────────────────
# Sheet type detection (unchanged logic)
# ──────────────────────────────────────────────
def _is_product_sheet(headers: List[str], max_col: int) -> bool:
    h_set = set(h.lower() for h in headers if h)
    has_asin = any("asin" in h for h in h_set)
    has_sales = any(kw in h for h in h_set for kw in ["月销量", "monthly sales", "月销售额"])
    has_price = any(kw in h for h in h_set for kw in ["价格", "price"])
    return has_asin and (has_sales or has_price) and max_col > 15


def _is_review_sheet(headers: List[str]) -> bool:
    h_set = set(h.lower() for h in headers if h)
    has_content = any(kw in h for h in h_set for kw in ["内容", "content", "body"])
    has_rating = any(kw in h for h in h_set for kw in ["星级", "rating"])
    has_asin = any("asin" in h for h in h_set)
    has_sales = any(kw in h for h in h_set for kw in ["月销量", "monthly sales"])
    return has_asin and has_content and has_rating and not has_sales


def _is_market_sheet(headers: List[str]) -> bool:
    ht = _header_text(headers)
    has_agg = any(kw in ht for kw in ["月总销量", "月均销量", "月均销售额", "月总销售额"])
    has_avg = any(kw in ht for kw in ["平均价格", "平均星级", "平均评分"])
    has_asin = any("asin" in h.lower() for h in headers if h)
    return (has_agg or has_avg) and not has_asin


def _is_competitor_sheet(headers: List[str]) -> bool:
    # Pure ASIN column headers (e.g. "B0CYWSWZ71")
    asin_count = sum(1 for h in headers if re.match(r"^[A-Z0-9]{10}$", h))
    if asin_count >= 2:
        return True
    # Embedded ASINs in complex headers (e.g. "竞品1: Zetunlo (B0CYWSWZ71)\n...")
    embedded_count = sum(1 for h in headers if h and re.search(r"[A-Z0-9]{10}", h)
                         and any(kw in h for kw in ("竞品", "Competitor", "competitor")))
    return embedded_count >= 2


def _is_profit_sheet(headers: List[str]) -> bool:
    ht = _header_text(headers)
    cost_kws = ["毛利", "fba", "头程运费", "产品成本", "毛利润", "fba总成本", "fba配送费", "fba佣金"]
    return sum(1 for kw in cost_kws if kw in ht) >= 2


def _is_keyword_sheet(headers: List[str]) -> bool:
    ht = _header_text(headers)
    has_kw = any(kw in ht for kw in ["关键词", "keyword"])
    has_metric = any(kw in ht for kw in ["搜索量", "月搜索量", "cpc", "竞价", "转化", "点击"])
    return has_kw and has_metric


def _is_trend_sheet(sheet_name: str, max_row: int) -> bool:
    name_l = sheet_name.lower()
    return max_row <= 15 and ("趋势" in name_l or "trend" in name_l)


def _is_conclusion_sheet(sheet_name: str, max_col: int) -> bool:
    name_l = sheet_name.lower()
    return max_col <= 6 and any(kw in name_l for kw in ["结论", "建议", "总结", "conclusion"])


# ──────────────────────────────────────────────
# Parsing functions (all use iter_rows)
# ──────────────────────────────────────────────
_NUMERIC_FIELDS = frozenset(("price", "monthly_sales", "monthly_revenue", "rating",
                              "ratings_total", "main_bsr", "sub_bsr", "fba_fee", "margin"))


def _parse_product_rows(ws, header_row: int, headers: List[str]) -> List[Dict[str, Any]]:
    norm = [_normalise_header(h) for h in headers]

    # ── Pre-row gate: targeted error when a required column is unmappable ──
    # This catches the "alias missing in _COLUMN_MAP" bug class with a clear,
    # actionable message instead of N rows of generic pydantic errors.
    if "title" not in norm:
        raw_headers = [str(h) for h in headers if h]
        raise ValueError(
            f"产品 sheet 「{ws.title}」未能识别标题列。"
            f"请检查表头是否包含「商品标题」「产品标题（完整）」「Product Title」之一，"
            f"或在 backend/excel_parser/parser.py 的 _COLUMN_MAP 中补充别名。\n"
            f"当前表头: {raw_headers[:15]}"
        )
    if "asin" not in norm:
        raw_headers = [str(h) for h in headers if h]
        raise ValueError(
            f"产品 sheet 「{ws.title}」未能识别 ASIN 列。"
            f"请检查表头是否包含「ASIN」列。\n当前表头: {raw_headers[:15]}"
        )

    ncols = len(norm)
    products: List[Dict[str, Any]] = []
    errors: List[Tuple[str, str]] = []  # (asin, error message)

    for row in ws.iter_rows(min_row=header_row + 1, max_col=ncols, values_only=True):
        row_dict: Dict[str, Any] = {}
        for i in range(min(len(row), ncols)):
            key = norm[i]
            if key and row[i] is not None:
                row_dict[key] = row[i]
        asin = row_dict.get("asin")
        if not asin or not str(asin).strip():
            continue
        row_dict["asin"] = str(asin).strip()
        for field in _NUMERIC_FIELDS:
            v = row_dict.get(field)
            if v is not None:
                f = _safe_float(v)
                if f is not None:
                    row_dict[field] = f

        # ── Validate against Product schema (parse-time contract gate) ──
        try:
            Product.model_validate(row_dict)
        except ValidationError as e:
            msgs = "; ".join(
                f"{'.'.join(str(p) for p in err['loc'])}: {err['msg']}"
                for err in e.errors()
            )
            errors.append((row_dict["asin"], msgs))
            continue  # skip invalid rows

        products.append(row_dict)

    # ── Post-loop: hard-fail if every row was invalid; warn if partial ──
    if errors and not products:
        sample = "\n".join(f"  - {a}: {m}" for a, m in errors[:10])
        raise ValueError(
            f"产品 sheet 「{ws.title}」校验失败：{len(errors)} 条记录全部不符合 Product schema。\n"
            f"示例错误:\n{sample}"
        )
    if errors:
        print(f"[parser] WARN: sheet 「{ws.title}」 跳过了 {len(errors)} 条无效产品行")
        for asin, msg in errors[:5]:
            print(f"  - {asin}: {msg}")

    return products


def _parse_review_rows(ws, header_row: int, headers: List[str]) -> List[Dict[str, Any]]:
    norm = [_normalise_review_header(h) for h in headers]
    ncols = len(norm)
    reviews = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=ncols, values_only=True):
        row_dict: Dict[str, Any] = {}
        for i in range(min(len(row), ncols)):
            key = norm[i]
            if key and row[i] is not None:
                row_dict[key] = row[i]
        body = row_dict.get("body")
        if not body or not str(body).strip():
            continue
        reviews.append({
            "id": row_dict.get("review_url", ""),
            "title": row_dict.get("title", ""),
            "body": str(body).strip(),
            "rating": _safe_float(row_dict.get("rating")),
            "date": None,
            "verified_purchase": str(row_dict.get("verified_purchase", "")).upper() in ("Y", "YES", "TRUE", "1"),
        })
    return reviews


def _parse_generic_rows(ws, header_row: int, headers: List[str], max_rows: int = 0) -> List[Dict[str, Any]]:
    """Generic row parser for market/profit/keyword sheets."""
    ncols = len(headers)
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=ncols, values_only=True):
        row_dict = {}
        for i in range(min(len(row), ncols)):
            if headers[i] and row[i] is not None:
                row_dict[headers[i]] = row[i]
        if row_dict:
            rows.append(row_dict)
        if max_rows and len(rows) >= max_rows:
            break
    return rows


def _parse_competitor_sheet(ws, header_row: int, headers: List[str]) -> Dict[str, Dict[str, Any]]:
    """Parse competitor matrix — ASINs as column headers, metrics as rows."""
    asins = []
    asin_start = 0
    for i, h in enumerate(headers):
        if re.match(r"^[A-Z0-9]{10}$", h):
            if not asins:
                asin_start = i
            asins.append(h)
        elif h:
            m = re.search(r"\b([A-Z0-9]{10})\b", h)
            if m:
                if not asins:
                    asin_start = i
                asins.append(m.group(1))
    if not asins:
        return {}

    result: Dict[str, Dict[str, Any]] = {a: {} for a in asins}
    for row in ws.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True):
        label = None
        extra_texts = []  # Additional text values found in label columns
        for i in range(min(asin_start, len(row))):
            if row[i] is not None and _safe_str(row[i]):
                s = _safe_str(row[i])
                if s.startswith("=DISPIMG"):
                    continue
                if label is None:
                    label = s
                else:
                    extra_texts.append(s)
        if not label:
            continue
        # Store per-ASIN values
        has_asin_data = False
        for j, asin in enumerate(asins):
            col_idx = asin_start + j
            if col_idx < len(row) and row[col_idx] is not None:
                result[asin][label] = row[col_idx]
                has_asin_data = True
        # If no ASIN column has data but there's meaningful text in label area,
        # store it for all ASINs (e.g., 好评总结, 差评总结, 使用场景, 用户画像)
        if not has_asin_data and extra_texts:
            longest = max(extra_texts, key=len)
            if len(longest) > 10:
                for asin in asins:
                    result[asin][label] = longest
    return result


_PROFIT_PARAM_MAP: Dict[str, str] = {
    "目标售价($)": "售价USD",
    "目标售价(usd)": "售价USD",
    "售价($)": "售价USD",
    "售价(usd)": "售价USD",
    "产品成本(cny)": "产品成本CNY",
    "产品成本(rmb)": "产品成本CNY",
    "产品成本(usd)": "产品成本USD",
    "头程运费(usd)": "头程运费CNY",  # map to CNY key for compatibility
    "头程运费(cny)": "头程运费CNY",
    "fba费用(usd)": "FBA总成本USD",
    "fba费用": "FBA总成本USD",
    "fba总成本(usd)": "FBA总成本USD",
    "amazon佣金": "FBA佣金USD",
    "广告费率": "广告费率",
    "广告费(usd)": "广告USD",
    "退货率": "退货率",
    "退货成本(usd)": "退货成本USD",
    "其他费用": "其他费用",
    "总成本(usd)": "总成本USD",
    "毛利润(usd)": "毛利润",
    "毛利率": "毛利率",
    "月销量目标": "月销量目标",
    "月毛利润(usd)": "月毛利润USD",
    "年毛利润(usd)": "年毛利润USD",
}


def _parse_transposed_profit(ws) -> List[Dict[str, Any]]:
    """Parse transposed profit sheet (rows=params, cols=schemes).

    E.g.:
        Row2: ['参数',       '低价方案', '主流方案', '中高端方案']
        Row3: ['目标售价($)', 29.99,     59.99,     69.99]
    Returns one dict per scheme column.
    """
    # Find the header row with scheme names
    scheme_row = None
    scheme_names = []
    param_col = 0
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=False):
        cells = [c.value for c in row]
        strs = [str(c).strip() if c is not None else "" for c in cells]
        # Look for the row with "参数" or the one with multiple scheme labels
        if any(kw in strs[0].lower() if strs else False for kw in ["参数", "指标", "项目"]):
            scheme_row = row[0].row
            param_col = 0
            scheme_names = [s for s in strs[1:] if s]
            break
        # Also detect by having 3+ non-empty cols with meaningful names
        non_empty = [s for s in strs if s and len(s) > 1]
        if len(non_empty) >= 3 and any(kw in " ".join(strs).lower() for kw in ["方案", "plan"]):
            scheme_row = row[0].row
            param_col = 0
            scheme_names = [s for s in strs[1:] if s]
            break

    if not scheme_row or not scheme_names:
        return []

    # Read parameter rows
    num_schemes = len(scheme_names)
    result = [{} for _ in range(num_schemes)]
    for i, name in enumerate(scheme_names):
        result[i]["运输方式"] = name  # Use as label for pricing_node

    for row in ws.iter_rows(min_row=scheme_row + 1, values_only=True):
        if not row or row[param_col] is None:
            continue
        param_name = str(row[param_col]).strip().lower()
        mapped = _PROFIT_PARAM_MAP.get(param_name)
        if not mapped:
            # Try partial match
            for pat, field in _PROFIT_PARAM_MAP.items():
                if pat in param_name:
                    mapped = field
                    break
        if not mapped:
            mapped = str(row[param_col]).strip()

        for i in range(num_schemes):
            col_idx = 1 + i
            if col_idx < len(row) and row[col_idx] is not None:
                result[i][mapped] = row[col_idx]

    return result


def _parse_text_sheet(ws, max_rows: int = 200) -> List[str]:
    """Parse a text-heavy sheet — collect all non-empty cell values."""
    texts = []
    count = 0
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                s = _safe_str(cell)
                if s and len(s) > 5 and not s.startswith("=DISPIMG"):
                    texts.append(s)
        count += 1
        if count >= max_rows:
            break
    return texts


# ──────────────────────────────────────────────
# ASIN extraction for review sheets
# ──────────────────────────────────────────────
def _extract_review_asin(sheet_name: str, ws, header_row: int, headers: List[str], file_path: str) -> Optional[str]:
    m = re.match(r"([A-Z0-9]{10})", sheet_name)
    if m:
        return m.group(1)
    # Check first data rows for ASIN column
    norm = [_normalise_review_header(h) for h in headers]
    asin_col = None
    for i, key in enumerate(norm):
        if key == "asin":
            asin_col = i
            break
    if asin_col is not None:
        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True):
            if asin_col < len(row) and row[asin_col]:
                return str(row[asin_col]).strip()
            count += 1
            if count >= 2:
                break
    fname_match = re.search(r"([A-Z0-9]{10})-", file_path.replace("\\", "/").split("/")[-1])
    if fname_match:
        return fname_match.group(1)
    return None


# ──────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────
def _empty_result() -> Dict[str, Any]:
    return {
        "products": {}, "reviews": {},
        "market_analysis": [], "competitor_analysis": {},
        "profit_calc": [], "keyword_analysis": [],
        "category_trends": [], "conclusions": [],
        "file_type": "unknown", "sheets_parsed": [],
    }


def classify_and_parse_file(file_path: str) -> Dict[str, Any]:
    """Parse a single Excel file, auto-detecting all sheet types."""
    fname = file_path.replace("\\", "/").split("/")[-1]
    if fname.startswith("~$"):
        return _empty_result()

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"[excel_parser] Cannot open {file_path}: {e}")
        return _empty_result()

    products_by_asin: Dict[str, Dict[str, Any]] = {}
    reviews_by_asin: Dict[str, List[Dict[str, Any]]] = {}
    market_analysis: List[Dict[str, Any]] = []
    competitor_analysis: Dict[str, Dict[str, Any]] = {}
    profit_calc: List[Dict[str, Any]] = []
    keyword_analysis: List[Dict[str, Any]] = []
    category_trends: List[str] = []
    conclusions: List[str] = []
    sheets_info = []
    found_product = False
    found_review = False

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ("note", "说明", "brands", "sellers"):
            continue
        ws = wb[sheet_name]
        max_col = ws.max_column or 0
        max_row = ws.max_row or 0
        header_row, headers = _get_headers(ws)

        # ── Name-based detection for sheets that header-based detection misses ──
        name_l = sheet_name.lower().replace(" ", "")
        if any(kw in name_l for kw in ["市场分析", "市场规模", "marketanalysis"]) and not headers:
            texts = _parse_text_sheet(ws, max_rows=MAX_ANALYSIS_ROWS)
            if texts:
                market_analysis.extend(texts)
                sheets_info.append({"name": sheet_name, "type": "market_text", "rows": len(texts)})
            continue

        if any(kw in name_l for kw in ["利润测算", "利润核算", "profitcalc"]) and not _is_profit_sheet(headers):
            data = _parse_transposed_profit(ws)
            if data:
                profit_calc.extend(data)
                sheets_info.append({"name": sheet_name, "type": "profit_transposed", "rows": len(data)})
                continue

        if _is_review_sheet(headers):
            found_review = True
            reviews = _parse_review_rows(ws, header_row, headers)
            if reviews:
                asin = _extract_review_asin(sheet_name, ws, header_row, headers, file_path)
                if asin:
                    reviews_by_asin.setdefault(asin, []).extend(reviews)
                    sheets_info.append({"name": sheet_name, "type": "review", "rows": len(reviews), "asin": asin})

        elif _is_product_sheet(headers, max_col):
            found_product = True
            rows = _parse_product_rows(ws, header_row, headers)
            for p in rows:
                a = p.get("asin")
                if a:
                    products_by_asin[a] = p
            sheets_info.append({"name": sheet_name, "type": "product", "rows": len(rows)})

        elif _is_competitor_sheet(headers):
            data = _parse_competitor_sheet(ws, header_row, headers)
            competitor_analysis.update(data)
            sheets_info.append({"name": sheet_name, "type": "competitor", "rows": len(data)})

        elif _is_market_sheet(headers):
            data = _parse_generic_rows(ws, header_row, headers, max_rows=MAX_ANALYSIS_ROWS)
            market_analysis.extend(data)
            sheets_info.append({"name": sheet_name, "type": "market", "rows": len(data)})

        elif _is_keyword_sheet(headers):
            data = _parse_generic_rows(ws, header_row, headers, max_rows=MAX_ANALYSIS_ROWS)
            keyword_analysis.extend(data)
            sheets_info.append({"name": sheet_name, "type": "keyword", "rows": len(data)})

        elif _is_profit_sheet(headers):
            data = _parse_generic_rows(ws, header_row, headers, max_rows=100)
            profit_calc.extend(data)
            sheets_info.append({"name": sheet_name, "type": "profit", "rows": len(data)})

        elif _is_trend_sheet(sheet_name, max_row):
            texts = _parse_text_sheet(ws, max_rows=20)
            category_trends.extend(texts)
            sheets_info.append({"name": sheet_name, "type": "trend", "rows": len(texts)})

        elif _is_conclusion_sheet(sheet_name, max_col):
            texts = _parse_text_sheet(ws, max_rows=100)
            conclusions.extend(texts)
            sheets_info.append({"name": sheet_name, "type": "conclusion", "rows": len(texts)})

        else:
            name_l = sheet_name.lower()
            if any(kw in name_l for kw in ["结论", "建议", "总结"]):
                texts = _parse_text_sheet(ws, max_rows=100)
                conclusions.extend(texts)
                sheets_info.append({"name": sheet_name, "type": "conclusion", "rows": len(texts)})
            elif any(kw in name_l for kw in ["趋势", "trend"]):
                texts = _parse_text_sheet(ws, max_rows=20)
                category_trends.extend(texts)
                sheets_info.append({"name": sheet_name, "type": "trend", "rows": len(texts)})
            elif any(kw in name_l for kw in ["关键词", "keyword"]):
                texts = _parse_text_sheet(ws, max_rows=100)
                keyword_analysis.extend({"text": t} for t in texts)
                sheets_info.append({"name": sheet_name, "type": "keyword_text", "rows": len(texts)})
            else:
                sheets_info.append({"name": sheet_name, "type": "skipped", "rows": 0})

    wb.close()

    # ── Extract and parse embedded xlsx files (e.g. review exports) ──
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            embedded = [n for n in zf.namelist()
                        if n.startswith("xl/embeddings/") and n.endswith(".xlsx")]
            if embedded:
                tmp_dir = tempfile.mkdtemp(prefix="xlsx_embed_")
                for name in embedded:
                    tmp_path = os.path.join(tmp_dir, os.path.basename(name))
                    with open(tmp_path, "wb") as f:
                        f.write(zf.read(name))
                    try:
                        sub = classify_and_parse_file(tmp_path)
                        for asin, revs in sub["reviews"].items():
                            reviews_by_asin.setdefault(asin, []).extend(revs)
                            if revs:
                                found_review = True
                        for asin, prod in sub["products"].items():
                            products_by_asin.setdefault(asin, prod)
                        if sub["reviews"]:
                            sheets_info.append({
                                "name": os.path.basename(name), "type": "embedded_review",
                                "rows": sum(len(v) for v in sub["reviews"].values()),
                            })
                    finally:
                        os.unlink(tmp_path)
                os.rmdir(tmp_dir)
    except (zipfile.BadZipFile, OSError):
        pass  # Not a valid zip or can't read — skip silently

    file_type = "unknown"
    if found_product and found_review:
        file_type = "mixed"
    elif found_product:
        file_type = "product"
    elif found_review:
        file_type = "review"

    return {
        "products": products_by_asin,
        "reviews": reviews_by_asin,
        "market_analysis": market_analysis,
        "competitor_analysis": competitor_analysis,
        "profit_calc": profit_calc,
        "keyword_analysis": keyword_analysis,
        "category_trends": category_trends,
        "conclusions": conclusions,
        "file_type": file_type,
        "sheets_parsed": sheets_info,
    }


def parse_all_files(file_paths: List[str]) -> Dict[str, Any]:
    """Parse multiple Excel files in parallel, merge all data types."""
    merged_products: Dict[str, Dict[str, Any]] = {}
    merged_reviews: Dict[str, List[Dict[str, Any]]] = {}
    merged_market: List[Dict[str, Any]] = []
    merged_competitor: Dict[str, Dict[str, Any]] = {}
    merged_profit: List[Dict[str, Any]] = []
    merged_keyword: List[Dict[str, Any]] = []
    merged_trends: List[str] = []
    merged_conclusions: List[str] = []
    all_sheets = []
    product_file_count = 0
    review_file_count = 0

    # Parse files in parallel
    with ThreadPoolExecutor(max_workers=min(4, len(file_paths))) as pool:
        futures = {pool.submit(classify_and_parse_file, fp): fp for fp in file_paths}
        for future in as_completed(futures):
            result = future.result()
            merged_products.update(result["products"])
            for asin, revs in result["reviews"].items():
                merged_reviews.setdefault(asin, []).extend(revs)
            merged_market.extend(result["market_analysis"])
            merged_competitor.update(result["competitor_analysis"])
            merged_profit.extend(result["profit_calc"])
            merged_keyword.extend(result["keyword_analysis"])
            merged_trends.extend(result["category_trends"])
            merged_conclusions.extend(result["conclusions"])
            all_sheets.extend(result["sheets_parsed"])
            if result["file_type"] in ("product", "mixed"):
                product_file_count += 1
            if result["file_type"] in ("review", "mixed"):
                review_file_count += 1

    total_reviews = sum(len(v) for v in merged_reviews.values())
    return {
        "excel_data": merged_products,
        "reviews_by_asin": merged_reviews,
        "market_analysis": merged_market,
        "competitor_analysis": merged_competitor,
        "profit_calc": merged_profit,
        "keyword_analysis": merged_keyword,
        "category_trends": merged_trends,
        "source_conclusions": merged_conclusions,
        "stats": {
            "product_files": product_file_count,
            "review_files": review_file_count,
            "total_products": len(merged_products),
            "total_reviews": total_reviews,
        },
        "all_sheets": all_sheets,
    }
